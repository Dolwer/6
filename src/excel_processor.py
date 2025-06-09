import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
import shutil
import datetime
import logging
import json
import re
import traceback
from .utils import normalize_email, validate_excel_structure, extract_emails, safe_filename

class ExcelProcessor:
    def __init__(self, file_path, logger, mail_column="Mail", target_columns=None, backup=True):
        self.file_path = Path(file_path)
        self.logger = logger
        self.mail_column = mail_column
        self.target_columns = target_columns or []
        self.backup = backup
        self.df = None
        self.email_index = {}
        self.changes = set()  # (row_idx, col_name)
        self.excel_log_path = Path("logs") / "excel_updates.log"
        self.excel_logger = self._setup_excel_logger()
        self.load_data()

    def _setup_excel_logger(self):
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        excel_logger = logging.getLogger('ExcelUpdates')
        excel_logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler(self.excel_log_path, encoding='utf-8')
        formatter = logging.Formatter('[%(asctime)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        handler.setFormatter(formatter)
        if not excel_logger.handlers:
            excel_logger.addHandler(handler)
        return excel_logger

    def load_data(self):
        """
        Загрузка Excel файла.
        Проверка существования файла и необходимых колонок.
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        self.df = pd.read_excel(self.file_path, engine="openpyxl")
        validate_excel_structure(self.df, self.target_columns + [self.mail_column])
        self.create_email_index()
        self.logger.info(f"Loaded Excel: {self.file_path}, rows: {len(self.df)}")

    def create_email_index(self):
        """
        Индексирует каждый email отдельно (на случай, если в ячейке их несколько).
        """
        self.email_index = {}
        for idx, row in self.df.iterrows():
            all_emails = extract_emails(str(row[self.mail_column]))
            for em in all_emails:
                norm = normalize_email(em)
                if norm not in self.email_index:
                    self.email_index[norm] = []
                self.email_index[norm].append(idx)
        self.logger.info(f"Excel email index created. Unique emails: {len(self.email_index)}")
        self.excel_logger.info(f"Created email index: {list(self.email_index.keys())[:20]}...")

    def _extract_json_from_text(self, text):
        """
        Устойчивое извлечение JSON из текста LLM с множественными стратегиями.
        """
        if not text or not isinstance(text, str):
            return None

        # 1. Поиск между ```json ... ``` (markdown-LLM)
        json_block_pattern = r'```json\s*(\{.*?\})\s*```'
        matches = re.findall(json_block_pattern, text, re.DOTALL | re.IGNORECASE)
        for match in matches:
            try:
                parsed = json.loads(match.strip())
                if self._is_valid_response_json(parsed):
                    return parsed
            except json.JSONDecodeError:
                continue

        # 2. Поиск любых JSON c нужными полями (Price usd, Payment, Inform, ...)
        json_pattern = r'\{[^{}]*(?:"Price usd"[^{}]*|"Payment"[^{}]*|"Inform"[^{}]*)\}'
        matches = re.findall(json_pattern, text, re.DOTALL)
        for match in matches:
            try:
                parsed = json.loads(match.strip())
                if self._is_valid_response_json(parsed):
                    return parsed
            except json.JSONDecodeError:
                continue

        # 3. Построчный парсинг с балансировкой скобок (edge-case)
        return self._parse_json_line_by_line(text)

    def _parse_json_line_by_line(self, text):
        """
        Парсинг JSON построчно с балансировкой скобок и учётом строк.
        """
        lines = text.split('\n')
        current_json = ""
        brace_count = 0
        in_string = False
        escape_next = False
        for line in lines:
            for char in line:
                if escape_next:
                    current_json += char
                    escape_next = False
                    continue
                if char == '\\' and in_string:
                    current_json += char
                    escape_next = True
                    continue
                if char == '"' and not escape_next:
                    in_string = not in_string
                    current_json += char
                    continue
                if not in_string:
                    if char == '{':
                        if brace_count == 0:
                            current_json = ""
                        current_json += char
                        brace_count += 1
                    elif char == '}':
                        current_json += char
                        brace_count -= 1
                        if brace_count == 0:
                            try:
                                parsed = json.loads(current_json.strip())
                                if self._is_valid_response_json(parsed):
                                    return parsed
                            except json.JSONDecodeError:
                                pass
                            current_json = ""
                    elif brace_count > 0:
                        current_json += char
                else:
                    current_json += char
        return None

    def _is_valid_response_json(self, obj):
        """
        Проверяет, что JSON содержит хотя бы одно из ключевых полей с непустым значением.
        """
        if not isinstance(obj, dict):
            return False

        expected_fields = ["Price usd", "Price usd casino", "Payment", "Inform", "Comments"]
        # хотя бы одно поле из списка
        has_expected_field = any(field in obj for field in expected_fields)
        if not has_expected_field:
            return False
        # хотя бы одно непустое значение
        has_non_empty_value = any(
            value and str(value).strip()
            for value in (obj.get(field, "") for field in expected_fields)
        )
        return has_non_empty_value

    def parse_lm_studio_response(self, raw_response):
        """
        Устойчивый парсер ответа от LM Studio, поддержка markdown, разных API-форматов.
        """
        self.excel_logger.info(f"=== PARSING LM STUDIO RESPONSE ===")
        self.excel_logger.info(f"Raw response type: {type(raw_response)}")
        try:
            if raw_response is None:
                self.excel_logger.error("Raw response is None")
                return {}
            response_str = str(raw_response)
            self.excel_logger.info(f"Raw response length: {len(response_str)} characters")
            self.excel_logger.info(f"Raw response preview (first 500 chars): {response_str[:500]}")
        except Exception as e:
            self.excel_logger.error(f"Failed to convert raw_response to string: {e}")
            return {}

        # Извлекаем текст из API-ответа
        text_content = self._extract_text_from_api_response(response_str)
        # Пытаемся вытащить JSON
        parsed_json = self._extract_json_from_text(text_content)

        if parsed_json:
            self.excel_logger.info(f"=== PARSING SUCCESS ===")
            self.excel_logger.info(f"Successfully parsed JSON: {parsed_json}")
            return parsed_json
        else:
            self.excel_logger.error("=== PARSING FAILED ===")
            self.excel_logger.error("No valid JSON objects found in response")
            self.excel_logger.error(f"Text content was: {text_content[:1000]}")
            return {}

    def _extract_text_from_api_response(self, response_str):
        """
        Извлекает текстовое содержимое из API-ответа LM Studio/OpenAI.
        """
        try:
            if '"choices"' in response_str and '"text"' in response_str:
                api_response = json.loads(response_str)
                text_content = api_response['choices'][0]['text']
                self.excel_logger.info(f"Extracted text from API response. Length: {len(text_content)}")
                return text_content
            elif '"choices"' in response_str and '"message"' in response_str:
                api_response = json.loads(response_str)
                text_content = api_response['choices'][0]['message']['content']
                self.excel_logger.info(f"Extracted message content from API response. Length: {len(text_content)}")
                return text_content
        except (json.JSONDecodeError, KeyError, IndexError) as e:
            self.excel_logger.warning(f"Failed to parse as API response: {e}")
        # Если не API-ответ, возвращаем как есть
        return response_str

    def update_rows(self, email, raw_response):
        """
        Обновление строк в DataFrame для всех email-совпадений, с учётом парсинга ответа LLM.
        """
        self.excel_logger.info(f"=== UPDATE ROWS STARTED for email: {email} ===")
        try:
            extracted_data = self.parse_lm_studio_response(raw_response)
            if not extracted_data:
                self.logger.warning(f"No valid data extracted from LM Studio response for {email}")
                self.excel_logger.warning(f"No valid data extracted, skipping update for {email}")
                self.debug_email_search(email)
                return 0

            emails = extract_emails(email)
            if not emails:
                self.logger.warning(f"update_rows: No valid emails in input: {email}")
                self.excel_logger.warning(f"update_rows: No valid emails in input: {email}")
                return 0

            updated_any = False
            for em in emails:
                norm = normalize_email(em)
                self.excel_logger.info(f"Trying to update for normalized email: {norm}")
                if norm not in self.email_index:
                    self.logger.warning(f"Email not found in Excel: {norm}")
                    self.excel_logger.warning(f"Email not found in Excel index: {norm}")
                    self.excel_logger.info(f"Available emails in index (first 20): {list(self.email_index.keys())[:20]}")
                    continue

                update_count = 0
                for idx in self.email_index[norm]:
                    row_changes = []
                    self.excel_logger.info(f"Processing row {idx} for email {norm}")
                    for col in self.target_columns:
                        if col not in self.df.columns:
                            self.excel_logger.warning(f"Column '{col}' not found in DataFrame")
                            continue
                        old_val = str(self.df.at[idx, col]) if pd.notna(self.df.at[idx, col]) else ""
                        new_val = str(extracted_data.get(col, "")).strip()
                        self.excel_logger.debug(f"Column {col}: old='{old_val}', new='{new_val}'")
                        if new_val and old_val != new_val:
                            self.df.at[idx, col] = new_val
                            self.changes.add((idx, col))
                            row_changes.append((col, old_val, new_val))
                            update_count += 1
                            self.excel_logger.info(f"Updated {col}: '{old_val}' -> '{new_val}'")
                    if row_changes:
                        self.logger.info(f"Excel updated for {norm} at row {idx}: {row_changes}")
                        self.excel_logger.info(f"Row {idx} for {norm} updated: {row_changes}")
                        updated_any = True
                    else:
                        self.excel_logger.info(f"Row {idx} for {norm}: no changes needed")
                if not update_count:
                    self.excel_logger.info(f"Row(s) for {norm} found but nothing changed (data matched or empty update).")
            self.excel_logger.info(f"=== UPDATE ROWS COMPLETED ===")
            return int(updated_any)
        except Exception as e:
            self.excel_logger.error(f"Exception in update_rows: {str(e)}")
            self.excel_logger.error(f"Exception type: {type(e)}")
            self.excel_logger.error(f"Traceback: {traceback.format_exc()}")
            self.logger.error(f"Excel update failed for {email}: {str(e)}")
            return 0

    def save_with_backup(self):
        """
        Сохранение файла с резервной копией и подсветкой изменений.
        """
        if self.backup:
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = self.file_path.parent / f"{self.file_path.stem}_backup_{ts}{self.file_path.suffix}"
            shutil.copy2(self.file_path, backup_path)
            self.logger.info(f"Backup Excel created: {backup_path}")
            self.excel_logger.info(f"Backup Excel created: {backup_path}")

        self.df.to_excel(self.file_path, index=False, engine="openpyxl")
        if self.changes:
            self.highlight_changes()
            self.logger.info("Excel changes highlighted.")
            self.excel_logger.info("Excel changes highlighted.")

    def highlight_changes(self):
        """
        Подсветка изменённых ячеек в Excel файле (использует openpyxl).
        """
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for idx, col in self.changes:
            row_num = idx + 2  # pandas 0-based, Excel 1-based (+1 for header)
            if col in header:
                col_num = header.index(col) + 1
                ws.cell(row=row_num, column=col_num).fill = fill
        wb.save(self.file_path)
        self.logger.info(f"Highlighted {len(self.changes)} changed cells in Excel.")
        self.excel_logger.info(f"Highlighted {len(self.changes)} changed cells in Excel.")
        self.changes.clear()

    def debug_email_search(self, email):
        """
        Отладочный метод для проверки поиска email в индексе
        """
        self.excel_logger.info(f"=== DEBUG EMAIL SEARCH for: {email} ===")
        emails = extract_emails(email)
        self.excel_logger.info(f"Extracted emails: {emails}")
        for em in emails:
            norm = normalize_email(em)
            self.excel_logger.info(f"Normalized email: {norm}")
            if norm in self.email_index:
                rows = self.email_index[norm]
                self.excel_logger.info(f"Found in index at rows: {rows}")
                for row_idx in rows:
                    original_email = self.df.at[row_idx, self.mail_column]
                    self.excel_logger.info(f"Row {row_idx} original email: '{original_email}'")
            else:
                self.excel_logger.info(f"NOT found in index")
                similar = [k for k in self.email_index.keys() if norm.lower() in k.lower() or k.lower() in norm.lower()]
                if similar:
                    self.excel_logger.info(f"Similar emails in index: {similar[:10]}")
        self.excel_logger.info("=== END DEBUG EMAIL SEARCH ===")
