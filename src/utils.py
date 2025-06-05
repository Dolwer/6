import logging
from pathlib import Path
import pandas as pd
import re
import time
import random
from functools import wraps
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill

def setup_logging(log_level="INFO", log_file="bot.log"):
    """
    Настройка логирования с записью в файл и вывод в консоль
    """
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)

    logger = logging.getLogger('EmailBot')
    logger.setLevel(getattr(logging, log_level.upper()))
    logger.handlers.clear()

    formatter = logging.Formatter(
        '[%(asctime)s] [%(levelname)s] [%(name)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    file_handler = logging.FileHandler(
        log_dir / log_file, encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(getattr(logging, log_level.upper()))
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger

def extract_emails(text):
    """
    Извлекает все email-адреса из текста.
    """
    email_regex = r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+'
    return re.findall(email_regex, str(text))

def normalize_email(email):
    """
    Приводит email к нижнему регистру и убирает пробелы.
    """
    return str(email).strip().lower()

def mark_mail_column_multiple_emails(excel_path, mail_column="Mail", color="FFFF0000"):
    """
    Открывает Excel-файл и красит ячейки mail_column, если в них больше одной почты.
    color — hex-код (например, "FFFF0000" — красный).
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    # Получаем номер столбца по имени
    col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == mail_column:
            col_idx = idx
            break
    if col_idx is None:
        return
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        emails = extract_emails(cell.value)
        if len(emails) > 1:
            cell.fill = fill
    wb.save(excel_path)

def filter_rows_with_valid_mail(df, mail_column="Mail"):
    """
    Оставляет только строки, где:
    - В mail_column содержится ровно одна почта (строго одна!).
    - В остальных колонках не ищутся почты.
    Возвращает отфильтрованный DataFrame и Series с индексами строк, где было больше одной почты (для окраски).
    """
    def mail_checker(val):
        emails = extract_emails(val)
        return len(emails) == 1

    mask_mail_exists = df[mail_column].apply(lambda x: bool(extract_emails(x)) if pd.notna(x) else False)
    mask_mail_single = df[mail_column].apply(lambda x: len(extract_emails(x)) == 1 if pd.notna(x) else False)
    mask_mail_multi = df[mail_column].apply(lambda x: len(extract_emails(x)) > 1 if pd.notna(x) else False)
    df_valid = df[mask_mail_exists & mask_mail_single].copy()
    idx_multi = df[mask_mail_multi].index
    return df_valid, idx_multi

def find_duplicate_emails(df, mail_column="Mail"):
    """
    Возвращает Series с дублирующимися email (нормализованными).
    """
    emails = df[mail_column].apply(lambda x: normalize_email(extract_emails(x)[0]) if extract_emails(x) else "")
    return emails[emails.duplicated(keep=False)]

def validate_excel_structure(df, required_columns, logger=None):
    """
    Проверка структуры Excel файла:
    - Наличие обязательных колонок
    - (Дубликаты теперь не вызывают ошибку, только логируются)
    """
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing columns: {missing_columns}")

    mail_col = [col for col in required_columns if 'mail' in col.lower()]
    if mail_col:
        mail_col = mail_col[0]
        emails = df[mail_col]
        all_emails = []
        for e in emails:
            if pd.isna(e) or (isinstance(e, str) and not e.strip()):
                continue
            found_emails = extract_emails(e)
            if len(found_emails) == 1:
                all_emails.append(normalize_email(found_emails[0]))
        s = pd.Series(all_emails)
        dups = s[s.duplicated(keep=False)]
        if not dups.empty and logger:
            logger.warning(f"Duplicate emails found in the Mail column (will use only the first occurrence): {list(dups.unique())}")
    return True

def retry_with_backoff(max_retries=3, base_delay=1, backoff_factor=2):
    """
    Декоратор для повторных попыток с экспоненциальной задержкой.
    Использование:
    @retry_with_backoff(max_retries=3)
    def my_func(...):
        ...
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if attempt == max_retries - 1:
                        raise e
                    delay = base_delay * (backoff_factor ** attempt) + random.uniform(0, 1)
                    print(f"Retry {attempt + 1}/{max_retries} for {func.__name__} after {delay:.2f}s due to: {e}")
                    time.sleep(delay)
            return None
        return wrapper
    return decorator

def normalize_subject(subject):
    """
    Нормализация темы письма: убирает RE:, FWD: и прочий мусор, приводит к нижнему регистру.
    """
    if not subject:
        return ""
    pattern = r'^(RE(\[\d+\])?:|FWD?:|\[EXTERNAL\])\s*'
    while True:
        new_subject = re.sub(pattern, '', subject, flags=re.IGNORECASE).strip()
        if new_subject == subject:
            break
        subject = new_subject
    return subject.lower()

def strip_html_tags(html):
    """
    Убирает HTML-теги из строки и возвращает только текст.
    """
    if not html:
        return ""
    return BeautifulSoup(str(html), "html.parser").get_text()

def safe_filename(filename, max_length=255):
    """
    Делает строку безопасной для использования в качестве имени файла.
    Обрезает слишком длинные имена, убирает запрещённые символы.
    """
    # Убираем недопустимые символы для Windows, Linux, macOS
    filename = re.sub(r'[<>:"/\\|?*\n\r\t]', '_', str(filename))
    # Обрезаем имя файла, если оно слишком длинное (оставляем расширение)
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
        ext = '.' + ext
    else:
        name, ext = filename, ''
    name = name[:max_length - len(ext)]
    return f"{name}{ext}"
